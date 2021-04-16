
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt

wb = Workbook()

# criando as tabelas do aquivo excel
ws1 = wb.create_sheet('sheet_1')
ws2 = wb.create_sheet('sheet_2')

# inserindo dados -> (linha, coluna, valor)
ws1.cell(1, 1, float(10))
ws1.cell(2, 1, float(20))
ws1.cell(3, 1, float(30))
ws1.cell(4, 1, float(40))

ws2.cell(1, 1, float(15))
ws2.cell(2, 1, float(25))
ws2.cell(3, 1, float(35))
ws2.cell(4, 1, float(45))

# guardando os valores da coluna A
x = [ x.value for x in ws1['A'] ]
y = [ x.value for x in ws2['A'] ]

plt.plot(x, y)
plt.savefig('fig1.png')
plt.close()

img = openpyxl.drawing.image.Image('fig1.png')
wb['sheet_1'].add_image(img, 'C1')
# ws1.add_image('fig1.png')

wb.__delitem__('Sheet')
wb.save(filename='arquivo.xlsx')
wb.close()

